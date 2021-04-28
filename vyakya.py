import json
import datetime
import openpyxl
from openpyxl.workbook import Workbook
import pandas as pd
xlpath = "out.xlsx"
workbook=openpyxl.load_workbook(xlpath)
sheet=workbook.active


# wb = Workbook()

# ws1 = wb.create_sheet("Sheet")
# ws1.title = "Title_A"

# ws2 = wb.create_sheet("Sheet_c", 0)
# ws2.title = "Title_B"

# wb.save(filename = 'sample_book.xlsx')

with open("task_input_list.json") as blog_file:
    data = json.load(blog_file)
j=1
s=0

for i in range(77,133):
    print(data[i])#3
    k=i+2
    try:
        datetime.datetime.strptime(data[i], '%m/%d/%y')
        j=j+1

        m=1
        description=data[i+1]
        while m==1:
            adder=data[k]
            check=adder[0:2]
            if check.isalpha()==True:
                description=description+adder
                k=k+1

            else:
                m=0

                amount=data[k]

        print("y"+description)
        print("z"+amount)

        # if amount > 0:

        dateSeperated = data[i].split("/")
        print(dateSeperated)
        sheet.cell(row=j,column=1).value= data[i]
        sheet.cell(row=j,column=2).value= dateSeperated[1]
        sheet.cell(row=j,column=3).value= dateSeperated[0]
        sheet.cell(row=j,column=4).value= dateSeperated[2]
        sheet.cell(row=j,column=5).value= description
        sheet.cell(row=j,column=6).value= amount
        workbook.save(xlpath)
        print("True")
        # else:
        #     ws1 = workbook.active
        #     dateSeperated = data[i].split("/")
        #     print(dateSeperated)
        #     ws1.cell(row=j,column=1).value= data[i]
        #     ws1.cell(row=j,column=2).value= dateSeperated[1]
        #     ws1.cell(row=j,column=3).value= dateSeperated[0]
        #     ws1.cell(row=j,column=4).value= dateSeperated[2]
        #     ws1.cell(row=j,column=5).value= description
        #     ws1.cell(row=j,column=6).value= amount
        #     workbook.save(xlpath)
        #     print("True")



    except:
        print("Incorrect data format, should be YYYY-MM-DD")